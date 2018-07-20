#################################################################
#
#   9/8/16 - Python Script to update web scrape file
#
#   Jeff Zinkerman
#
#   Import CSV File from Web scrape.  
#   Modify column contents
#       replace K in miles with 000
#       set the ZIP to be the last 5 digits of the Zip=XXXXX in the URL field
#   Save modified file
#   Save to Excel ?
#################################################################
import csv
#import xlrd
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

def UpdateMileage(ml1):  
    # replace the K with a 000 to make a number
    ml1 = ml1.replace("K","000")    
    return ml1

def UpdateZIP(url1):
    # Update the ZIP Column
    # Take a long URL and Pull the zip off the back
    if len(url1) > 5 :
        tempstr = url1[-5:]
    else:
        tempstr = ''
    return tempstr

def UpdateURL(url1):
    if len(url1) > 50 :
        indx = url1.find("FreeText")
        if indx  != -1 :            
            tempstr = url1[indx+9:]
            tempstr = tempstr[:-10]
        else :
            tempstr = url1
    else :
        tempstr = url1
        
    return tempstr


def doCSVExtract():
    #open webscrape output file and write it to another output file in CSV
    file1= 'carmaxoutput1.csv'
    with open(file1) as csvfile1:
        reader1 = csv.DictReader(csvfile1)
        fieldnames = ['id','Date/Time', 'Price1', 'Mileage1', 'Price2', 'Mileage2', 'Price3', 'Mileage3', 'Price4', 'Mileage4', 'Price5', 'Mileage5','ZIP','URL']
        writer1 = csv.DictWriter( open('AllDataOutput.csv', 'w', newline=''),fieldnames=fieldnames, dialect='excel')
        writer1.writeheader()
        for row in reader1:
            # print(row['URL'])
            row['Mileage1'] = UpdateMileage(row['Mileage1'])
            row['Mileage2'] = UpdateMileage(row['Mileage2'])
            row['Mileage3'] = UpdateMileage(row['Mileage3'])
            row['Mileage4'] = UpdateMileage(row['Mileage4'])
            row['Mileage5'] = UpdateMileage(row['Mileage5'])
            row['ZIP'] = UpdateZIP(row['URL'])
            row['URL']= UpdateURL(row['URL'])

            writer1.writerow(row) 
            #writer1.writerow({'Date/Time' : row['Date/Time'], 'Price1' : row['Price1'] })
            
            print(row['ZIP'])
    

def doXLSXFile() :
    ## wkb1 = xlrd.open_workbook('AllDataOutput.xlsx')
    ## wks1 = wkb1.sheet_by_index(0)
    xlsx_output_file = 'AllDataOutput.xlsx'
    wbk1 = load_workbook(filename = xlsx_output_file)
    wks1 = wbk1.get_sheet_by_name(name = 'Analysis1')
    wks1['A200']= ('zzz')
    wbk1.save(filename = all_data_file)
    print('OK2')    

def main() :
    doXLSXFile()    
    # doCSVExtract()    
            
if __name__ == '__main__': main()
